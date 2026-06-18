# -*- coding: utf-8 -*-
"""tag_dictionary JSON 스냅샷 → 검토용 xlsx 생성."""
import json
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
SRC = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_tag_dict.json")
OUT = "D:/Dropbox/Claude Code/260503 피부텐텐 웹앱개발/전달용/260613 피부텐텐 태그 사전.xlsx"
KST = timezone(timedelta(hours=9))

with open(SRC, encoding="utf-8") as f:
    rows = json.load(f)

# ── 컬럼 정의 (헤더 한글, 원본 키, 너비) ──────────────────────────────
COLS = [
    ("id", "ID", 7),
    ("ko", "태그(한글)", 22),
    ("category", "분류", 12),
    ("en", "영문 slug", 24),
    ("parent_ko", "부모 태그", 16),
    ("is_procedure", "시술 여부", 9),
    ("is_recommendable", "추천 노출", 9),
    ("onboarding", "온보딩 분류", 14),
    ("sort_order", "정렬순서", 9),
    ("aliases", "동의어(별칭)", 30),
    ("pubmed_keywords", "PubMed 검색어", 30),
    ("reviewed_at", "검수일", 18),
    ("created_at", "생성일", 18),
    ("updated_at", "수정일", 18),
]

def fmt(key, val):
    if val is None:
        return ""
    if key in ("is_procedure", "is_recommendable"):
        return "O" if val else ""
    if isinstance(val, list):
        return ", ".join(str(x) for x in val)
    if key in ("reviewed_at", "created_at", "updated_at") and isinstance(val, str):
        try:
            dt = datetime.fromisoformat(val.replace("Z", "+00:00")).astimezone(KST)
            return dt.strftime("%Y-%m-%d %H:%M")
        except ValueError:
            return val
    return val

wb = Workbook()

# ══ Sheet 1: 태그 사전 ════════════════════════════════════════════════
ws = wb.active
ws.title = "태그 사전"

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(name="맑은 고딕", size=10)
PROC_FILL = PatternFill("solid", fgColor="FCE4D6")  # 시술 행 강조
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 헤더
for c, (_, label, width) in enumerate(COLS, start=1):
    cell = ws.cell(row=1, column=c, value=label)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(c)].width = width

# 데이터
for r, row in enumerate(rows, start=2):
    is_proc = row.get("is_procedure")
    for c, (key, _, _) in enumerate(COLS, start=1):
        cell = ws.cell(row=r, column=c, value=fmt(key, row.get(key)))
        cell.font = CELL_FONT
        cell.border = BORDER
        if key in ("id", "is_procedure", "is_recommendable", "sort_order"):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif key in ("aliases", "pubmed_keywords"):
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")
        if is_proc:
            cell.fill = PROC_FILL

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}{len(rows)+1}"

# ══ Sheet 2: 요약 ════════════════════════════════════════════════════
ws2 = wb.create_sheet("요약")
ws2.column_dimensions["A"].width = 20
ws2.column_dimensions["B"].width = 12

def put(r, a, b, header=False):
    ca = ws2.cell(row=r, column=1, value=a)
    cb = ws2.cell(row=r, column=2, value=b)
    if header:
        for cc in (ca, cb):
            cc.fill = HEADER_FILL
            cc.font = HEADER_FONT
            cc.alignment = Alignment(horizontal="center")
    else:
        ca.font = CELL_FONT
        cb.font = CELL_FONT
        cb.alignment = Alignment(horizontal="center")

now = datetime.now(KST).strftime("%Y-%m-%d %H:%M KST")
ws2.cell(row=1, column=1, value=f"피부텐텐 태그 사전 (tag_dictionary) — 추출 {now}").font = Font(
    name="맑은 고딕", bold=True, size=11)

put(3, "분류", "태그 수", header=True)
cat_count = {}
for row in rows:
    cat_count[row["category"]] = cat_count.get(row["category"], 0) + 1
r = 4
for cat, n in sorted(cat_count.items(), key=lambda x: -x[1]):
    put(r, cat, n)
    r += 1
put(r, "합계", len(rows))
r += 2
put(r, "시술(is_procedure)", sum(1 for x in rows if x.get("is_procedure"))); r += 1
put(r, "추천노출(is_recommendable)", sum(1 for x in rows if x.get("is_recommendable"))); r += 1
put(r, "부모태그 보유", sum(1 for x in rows if x.get("parent_ko"))); r += 1
put(r, "동의어 보유", sum(1 for x in rows if x.get("aliases"))); r += 1
put(r, "PubMed 검색어 보유", sum(1 for x in rows if x.get("pubmed_keywords"))); r += 1
put(r, "검수완료(reviewed_at)", sum(1 for x in rows if x.get("reviewed_at")))

wb.save(OUT)
print("OK:", OUT)
print("rows:", len(rows))
