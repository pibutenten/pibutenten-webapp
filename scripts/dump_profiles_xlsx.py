#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""members.xlsx 생성. 두 시트:
  - members: 회원 전체 + 모든 회원 변수
  - qa_author: 카드 글쓴이 관련 변수 정의 (참고용)
"""
import json, sys, urllib.request
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')
ROOT = Path(__file__).parent.parent
TOKEN = next(
    (l.split("=", 1)[1].strip()
     for l in (ROOT / ".env.local").read_text(encoding="utf-8").splitlines()
     if l.startswith("SUPABASE_ACCESS_TOKEN=")),
    None,
)
EP = "https://api.supabase.com/v1/projects/nahznfvouuwxqctwlwfs/database/query"
sql = """
SELECT
  u.email,
  p.display_name,
  p.handle,
  p.role,
  p.id          AS profile_id,
  p.auth_user_id,
  p.alt_handle,
  p.alt_display_name,
  p.bio,
  p.birth_date::text,
  p.gender,
  p.level,
  p.activity_score,
  p.is_public,
  p.terms_agreed_at::text,
  p.marketing_email_consent,
  d.id          AS doctor_id,
  d.slug        AS doctor_slug,
  d.name        AS doctor_name,
  d.branch      AS doctor_branch,
  p.created_at::text AS profile_created_at
FROM public.profiles p
LEFT JOIN auth.users u ON u.id = p.auth_user_id
LEFT JOIN public.doctor_accounts da ON da.profile_id = p.id
LEFT JOIN public.doctors d ON d.id = da.doctor_id
ORDER BY p.auth_user_id NULLS FIRST, p.created_at;
"""
req = urllib.request.Request(EP, data=json.dumps({"query": sql}).encode("utf-8"),
    method="POST", headers={
        "Authorization": f"Bearer {TOKEN}", "Content-Type": "application/json",
        "User-Agent": "pibutenten-dump/1.0",
    })
with urllib.request.urlopen(req, timeout=60) as resp:
    rows = json.loads(resp.read().decode('utf-8'))

wb = Workbook()
ws = wb.active
ws.title = "members"

cols = [
    "email", "display_name", "handle", "role",
    "profile_id", "auth_user_id", "alt_handle", "alt_display_name",
    "bio", "birth_date", "gender", "level", "activity_score",
    "is_public", "terms_agreed_at", "marketing_email_consent",
    "doctor_id", "doctor_slug", "doctor_name", "doctor_branch",
    "profile_created_at",
]
# header
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="2C5F7E")
for i, c in enumerate(cols, 1):
    cell = ws.cell(row=1, column=i, value=c)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
ws.freeze_panes = "A2"

# data
for r_i, r in enumerate(rows, 2):
    for c_i, c in enumerate(cols, 1):
        v = r.get(c)
        ws.cell(row=r_i, column=c_i, value="" if v is None else v)

# column widths
widths = {
    "email": 32, "display_name": 14, "handle": 18, "role": 8,
    "profile_id": 38, "auth_user_id": 38, "alt_handle": 14, "alt_display_name": 14,
    "bio": 30, "birth_date": 12, "gender": 8, "level": 6, "activity_score": 10,
    "is_public": 9, "terms_agreed_at": 22, "marketing_email_consent": 10,
    "doctor_id": 38, "doctor_slug": 16, "doctor_name": 12, "doctor_branch": 18,
    "profile_created_at": 22,
}
for i, c in enumerate(cols, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(c, 12)

# qa_author 정의 sheet
ws2 = wb.create_sheet("qa_author 변수")
qa_defs = [
    ("qas.author_id",  "profiles.id 참조", "그 글을 책임지는 profile (수정권한). Q&A 카드는 doctor의 profile.id"),
    ("qas.doctor_id",  "doctors.id 참조",  "그 카드의 화자 doctor entity. type=qa 카드 한정"),
    ("qas.posted_as",  "text",             "official / personal — 의사가 공식 명의 vs 개인 명의"),
    ("qas.hide_doctor_credential", "bool", "카드에서 '피부과 전문의' 직함 숨김 여부"),
    ("qas.type",       "text",             "qa / post / article"),
    ("qas.category",   "text",             "qa / tip / diary / ask / link"),
    ("comments.author_id", "auth.users.id 참조 (예외: profiles.id 아님)", "댓글 작성자. 마이그레이션 안 됨"),
]
ws2.cell(row=1, column=1, value="변수").font = header_font
ws2.cell(row=1, column=1).fill = header_fill
ws2.cell(row=1, column=2, value="타입").font = header_font
ws2.cell(row=1, column=2).fill = header_fill
ws2.cell(row=1, column=3, value="설명").font = header_font
ws2.cell(row=1, column=3).fill = header_fill
for i, (a, b, c) in enumerate(qa_defs, 2):
    ws2.cell(row=i, column=1, value=a)
    ws2.cell(row=i, column=2, value=b)
    ws2.cell(row=i, column=3, value=c)
ws2.column_dimensions["A"].width = 32
ws2.column_dimensions["B"].width = 30
ws2.column_dimensions["C"].width = 70
ws2.freeze_panes = "A2"

out = ROOT / "scripts" / "members.xlsx"
wb.save(out)
print(f"OK: {out} (members {len(rows)} rows + qa_author 변수 sheet)")
